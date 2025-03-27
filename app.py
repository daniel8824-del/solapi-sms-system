from flask import Flask, request, jsonify, render_template, send_from_directory, session
import os
from dotenv import load_dotenv
import json
import re
import requests
import boto3
from botocore.exceptions import NoCredentialsError
import uuid
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
import pandas as pd
import traceback
import tempfile

# .env 파일 로드
load_dotenv()
app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'solapi-secret-key-for-session')

# 데이터 폴더 정의
DATA_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
TEMP_UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')

# 데이터 폴더 생성
if not os.path.exists(DATA_FOLDER):
    os.makedirs(DATA_FOLDER)

# AWS Lambda 함수 URL (환경 변수로부터 로드)
LAMBDA_FUNCTION_URL = os.environ.get('LAMBDA_FUNCTION_URL', '')
print(f"Lambda Function URL: {LAMBDA_FUNCTION_URL}")

# 템플릿 다운로드 경로
BULK_TEMPLATE_PATH = os.path.join(DATA_FOLDER, 'sample_template.csv')
AUTO_TEMPLATE_PATH = os.path.join(DATA_FOLDER, 'automation_template.xlsx')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/send-sms', methods=['POST'])
def send_sms():
    """단일 메시지 발송 API - JSON 요청 처리"""
    try:
        # JSON 요청 데이터 가져오기
        data = request.get_json()
        to = data.get('to', '')
        message = data.get('message', '')
        image_data = data.get('image', None)
        
        print(f"단일 메시지 JSON 요청: to={to}, message={message[:20]}...")
        print(f"이미지 데이터: {json.dumps(image_data)}")
        
        if not to or not message:
            return jsonify({'success': False, 'message': '수신번호와 메시지 내용이 필요합니다.'}), 400
        
        # Lambda 요청 데이터 준비
        lambda_data = {
            'type': 'single',
            'to': to,
            'message': message
        }
        
        # 이미지가 있는 경우 처리
        if image_data:
            # 이미지 데이터 검증
            if isinstance(image_data, dict) and 'data' in image_data and image_data.get('data'):
                lambda_data['image'] = image_data
                print(f"이미지 첨부 요청 감지: {len(image_data['data'])}바이트")
            else:
                print(f"이미지 데이터가 비어있거나 잘못된 형식입니다.")
        
        # 디버깅 모드 활성화
        DEBUG_MODE = os.environ.get('DEBUG_MODE', 'True').lower() == 'true'
        
        # Lambda 함수 URL이 비어 있는 경우 또는 디버깅 모드일 경우 테스트 응답 반환
        if not LAMBDA_FUNCTION_URL or DEBUG_MODE:
            print("테스트 모드에서 실행 중입니다. 테스트 응답을 반환합니다.")
            return jsonify({
                'success': True,
                'message': '테스트 모드: 메시지가 성공적으로 발송된 것으로 처리됩니다.',
                'test_data': lambda_data
            })
        
        # Lambda 함수 호출
        print(f"Lambda 함수 호출: {LAMBDA_FUNCTION_URL}")
        response = requests.post(LAMBDA_FUNCTION_URL, json=lambda_data)
        print(f"Lambda 응답: status_code={response.status_code}, text={response.text[:100]}...")
        
        if response.status_code == 200:
            result = response.json()
            return jsonify(result)
        else:
            return jsonify({
                'success': False,
                'message': f'메시지 발송 실패: {response.text}'
            }), response.status_code
        
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/send-single', methods=['POST'])
def send_single():
    """단일 메시지 발송 API"""
    try:
        # 요청 데이터 가져오기
        to = request.form.get('to', '')
        message = request.form.get('message', '')
        image = request.files.get('image')
        
        print(f"단일 메시지 요청: to={to}, message={message[:20]}...")
        
        if not to or not message:
            return jsonify({'success': False, 'message': '수신번호와 메시지 내용이 필요합니다.'}), 400
        
        # Lambda 요청 데이터 준비
        lambda_data = {
            'type': 'single',
            'to': to,
            'message': message
        }
        
        # 이미지가 있는 경우 처리
        if image and image.filename:
            image_data = base64.b64encode(image.read()).decode('utf-8')
            lambda_data['image'] = {
                'data': image_data,
                'filename': image.filename
            }
        
        # 디버깅 모드 활성화
        DEBUG_MODE = os.environ.get('DEBUG_MODE', 'True').lower() == 'true'
        
        # Lambda 함수 URL이 비어 있는 경우 또는 디버깅 모드일 경우 테스트 응답 반환
        if not LAMBDA_FUNCTION_URL or DEBUG_MODE:
            print("테스트 모드에서 실행 중입니다. 테스트 응답을 반환합니다.")
            return jsonify({
                'success': True,
                'message': '테스트 모드: 메시지가 성공적으로 발송된 것으로 처리됩니다.',
                'test_data': lambda_data
            })
        
        # Lambda 함수 호출
        print(f"Lambda 함수 호출: {LAMBDA_FUNCTION_URL}")
        response = requests.post(LAMBDA_FUNCTION_URL, json=lambda_data)
        print(f"Lambda 응답: status_code={response.status_code}, text={response.text[:100]}...")
        
        if response.status_code == 200:
            result = response.json()
            return jsonify(result)
        else:
            return jsonify({
                'success': False,
                'message': f'메시지 발송 실패: {response.text}'
            }), response.status_code
        
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/send-bulk', methods=['POST'])
def send_bulk():
    """대량 메시지 발송 API (직접 입력)"""
    try:
        # 요청 데이터 가져오기
        text = request.form.get('text', '')
        recipients = request.form.get('recipients', '')
        recipient_list_json = request.form.get('recipientList', '')  # 추가: recipientList 파라미터 확인
        image = request.files.get('image')
        file = request.files.get('file')  # 파일 업로드 방식 확인
        
        # 디버깅을 위한 요청 내용 로깅
        print("=" * 50)
        print("대량 메시지 발송 요청 데이터:")
        print(f"text: {text[:30]}...")
        print(f"recipients: {recipients}")
        print(f"recipientList: {recipient_list_json}")
        print(f"file: {file.filename if file else 'None'}")
        for key, value in request.form.items():
            print(f"폼 데이터 - {key}: {value[:30] if len(str(value)) > 30 else value}...")
        print("=" * 50)
        
        # 파일 업로드 방식인 경우 CSV 파일 확인
        if file and file.filename:
            # 파일 확장자 확인
            file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
            if file_ext != 'csv':
                return jsonify({'success': False, 'message': 'CSV 형식의 파일만 지원합니다.'}), 400
            
            # 파일로부터 수신자 목록 추출
            response = parse_recipients_only_from_file(file, text)
            if not response.get('success', False):
                return jsonify(response), 400
                
            recipient_numbers = response.get('recipients', [])
        else:
            # 수신자 목록 처리 (recipientList가 있으면 우선 사용)
            recipient_numbers = []
            
            if recipient_list_json:
                try:
                    # JSON 형식으로 전달된 수신자 목록 처리
                    recipient_numbers = json.loads(recipient_list_json)
                    print(f"recipientList에서 수신자 목록 추출: {len(recipient_numbers)}명")
                except Exception as e:
                    print(f"recipientList JSON 파싱 오류: {str(e)}, 문자열로 처리합니다.")
                    recipient_numbers = recipient_list_json.split(',')
            
            # recipientList가 없으면 recipients 파라미터 사용
            if not recipient_numbers and recipients:
                # 쉼표 또는 줄바꿈으로 구분된 수신자 목록 처리
                if '\n' in recipients:
                    recipient_numbers = recipients.replace('\r', '').split('\n')
                else:
                    recipient_numbers = recipients.split(',')
                
                print(f"recipients에서 수신자 목록 추출: {len(recipient_numbers)}명")
            
            # 공백 제거 및 빈 항목 필터링
            recipient_numbers = [r.strip() for r in recipient_numbers if r.strip()]
            print(f"최종 수신자 수: {len(recipient_numbers)}명")
            print(f"최종 수신자 목록: {recipient_numbers}")
        
        if not text or not recipient_numbers:
            return jsonify({'success': False, 'message': '메시지 내용과 수신자 목록이 필요합니다.'}), 400
        
        # Lambda 요청 데이터 준비
        lambda_data = {
            'type': 'send_message',
            'text': text,
            'recipients': json.dumps(recipient_numbers)  # 리스트를 JSON 문자열로 변환
        }
        
        # 상세 로깅 추가
        print("=" * 80)
        print(f"Lambda로 전송하는 데이터:")
        print(f"text: {text[:30]}...")
        print(f"recipients(raw): {recipient_numbers}")
        print(f"recipients(JSON): {lambda_data['recipients']}")
        print(f"recipients JSON 길이: {len(lambda_data['recipients'])}")
        print(f"recipients JSON 예상 파싱 결과: {json.loads(lambda_data['recipients'])}")
        print("=" * 80)
        
        # 이미지가 있는 경우 처리
        if image and image.filename:
            image_data = base64.b64encode(image.read()).decode('utf-8')
            lambda_data['image'] = {
                'data': image_data,
                'filename': image.filename
            }
        
        # 디버깅 모드 활성화
        DEBUG_MODE = os.environ.get('DEBUG_MODE', 'True').lower() == 'true'
        
        # 디버깅 모드일 경우 테스트 응답 반환
        if DEBUG_MODE:
            print("테스트 모드에서 실행 중입니다. 대량 메시지 발송 테스트 응답을 반환합니다.")
            total_count = len(recipient_numbers)
            
            return jsonify({
                'success': True,
                'message': '테스트 모드: 메시지가 성공적으로 발송된 것으로 처리됩니다.',
                'total': total_count,
                'failedCount': 0
            })
        
        # Lambda 함수 호출
        print(f"Lambda 함수 호출: {LAMBDA_FUNCTION_URL}")
        response = requests.post(LAMBDA_FUNCTION_URL, json=lambda_data)
        print(f"Lambda 응답: status_code={response.status_code}, text={response.text}")
        
        if response.status_code == 200:
            result = response.json()
            print(f"Lambda 응답 파싱: {json.dumps(result, ensure_ascii=False)}")
            return jsonify(result)
        else:
            return jsonify({
                'success': False,
                'message': f'메시지 발송 실패: {response.text}'
            }), response.status_code
        
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/upload-excel', methods=['POST'])
def upload_excel():
    """엑셀 파일 업로드 및 메시지 미리보기 API"""
    try:
        # 파일 확인
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '파일이 필요합니다.'}), 400
            
        file = request.files['file']
        if not file or not file.filename:
            return jsonify({'success': False, 'message': '파일이 필요합니다.'}), 400
        
        # 디버깅 모드 활성화
        DEBUG_MODE = os.environ.get('DEBUG_MODE', 'True').lower() == 'true'
        
        # 디버깅 모드일 경우 테스트 응답 반환
        if DEBUG_MODE:
            print("테스트 모드에서 실행 중입니다. 엑셀 미리보기 테스트 응답을 반환합니다.")
            test_recipients = [
                {'to': '01012345678', 'text': '안녕하세요 김철수님, 2023-05-15에 주문하신 상품이 배송되었습니다.'},
                {'to': '01023456789', 'text': '안녕하세요 이영희님, 2023-05-16에 주문하신 상품이 배송되었습니다.'},
                {'to': '01034567890', 'text': '안녕하세요 박지민님, 2023-05-17에 주문하신 상품이 배송되었습니다.'},
                {'to': '01045678901', 'text': '안녕하세요 정민준님, 2023-05-18에 주문하신 상품이 배송되었습니다.'},
                {'to': '01056789012', 'text': '안녕하세요 윤서연님, 2023-05-19에 주문하신 상품이 배송되었습니다.'}
            ]
            return jsonify({
                'success': True,
                'recipients': test_recipients
            })
            
        # Lambda 요청 데이터 준비
        excel_data = base64.b64encode(file.read()).decode('utf-8')
        
        lambda_data = {
            'type': 'auto_excel_preview',
            'excel': {
                'data': excel_data,
                'filename': file.filename
            }
        }
        
        # Lambda 함수 호출
        response = requests.post(LAMBDA_FUNCTION_URL, json=lambda_data)
        
        if response.status_code == 200:
            result = response.json()
            
            # 결과 처리
            if result.get('success'):
                recipients = result.get('recipients', [])
                return jsonify({
                    'success': True,
                    'recipients': recipients
                })
            else:
                return jsonify({
                    'success': False,
                    'message': result.get('message', '엑셀 파일 처리 실패')
                }), 400
        else:
            return jsonify({
                'success': False,
                'message': f'엑셀 파일 처리 실패: {response.text}'
            }), response.status_code
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/send-excel', methods=['POST'])
def send_excel():
    """엑셀 기반 메시지 발송 API"""
    try:
        # 파일 확인
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '파일이 필요합니다.'}), 400
            
        file = request.files['file']
        image = request.files.get('image')
        
        if not file or not file.filename:
            return jsonify({'success': False, 'message': '파일이 필요합니다.'}), 400
            
        # 디버깅 모드 활성화
        DEBUG_MODE = os.environ.get('DEBUG_MODE', 'True').lower() == 'true'
        
        # 디버깅 모드일 경우 테스트 응답 반환
        if DEBUG_MODE:
            print("테스트 모드에서 실행 중입니다. 엑셀 기반 메시지 발송 테스트 응답을 반환합니다.")
            return jsonify({
                'success': True,
                'message': '테스트 모드: 메시지가 성공적으로 발송된 것으로 처리됩니다.',
                'total': 5,
                'failedCount': 0
            })
            
        # Lambda 요청 데이터 준비
        excel_data = base64.b64encode(file.read()).decode('utf-8')
        
        lambda_data = {
            'type': 'auto_excel_send',
            'excel': {
                'data': excel_data,
                'filename': file.filename
            }
        }
        
        # 이미지가 있는 경우 처리
        if image and image.filename:
            image_data = base64.b64encode(image.read()).decode('utf-8')
            lambda_data['image'] = {
                'data': image_data,
                'filename': image.filename
            }
        
        # Lambda 함수 호출
        response = requests.post(LAMBDA_FUNCTION_URL, json=lambda_data)
        
        if response.status_code == 200:
            result = response.json()
            return jsonify(result)
        else:
            return jsonify({
                'success': False,
                'message': f'메시지 발송 실패: {response.text}'
            }), response.status_code
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/download-template/<template_type>', methods=['GET'])
def download_template(template_type):
    """템플릿 파일 다운로드 API"""
    try:
        if template_type == 'bulk':
            # CSV 템플릿 다운로드 (기본 옵션)
            csv_template_path = os.path.join(DATA_FOLDER, 'sample_template.csv')
            
            # CSV 템플릿 파일이 없으면 생성
            if not os.path.exists(csv_template_path):
                # 간단한 CSV 템플릿 생성
                with open(csv_template_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(['이름', '휴대폰번호'])
                    writer.writerow(['홍길동', '010-1234-1234'])
                    writer.writerow(['전우치', '010-1234-1234'])
            
            return send_from_directory(DATA_FOLDER, 'sample_template.csv', as_attachment=True)
            
        elif template_type == 'auto':
            # 기존 템플릿 파일이 없으면 생성
            if not os.path.exists(AUTO_TEMPLATE_PATH):
                wb = Workbook()
                auto_sheet = wb.active
                auto_sheet.title = "자동메시지"
                
                # A1에 설명 추가
                auto_sheet['A1'] = "★ 아래 셀(A2)에 메시지 템플릿을 작성하세요. {{변수명}}으로 치환 가능합니다."
                
                # A2에 기본 템플릿 추가
                auto_sheet['A2'] = "[자동화 메시지 템플릿]\n안녕하세요 {{이름}}님, \n주문해주신 상품이 발송되었습니다.\n◎ 주문일자: {{주문일자}}\n◎ 주문금액: {{주문금액}}\n\n감사합니다."
                
                # 데이터 시트 만들기
                data_sheet = wb.create_sheet("data")
                
                # 헤더 추가
                headers = ["발송여부", "휴대폰번호", "이름", "주문일자", "주문금액", "주문상품"]
                for col, header in enumerate(headers, 1):
                    cell = data_sheet.cell(row=1, column=col)
                    cell.value = header
                
                # 샘플 데이터 추가
                sample_data = [
                    ["TRUE", "01012345678", "홍길동", "2025-03-22", "50,000원", "스마트폰 케이스"],
                    ["TRUE", "01098765432", "김철수", "2025-03-22", "35,000원", "블루투스 이어폰"],
                    ["FALSE", "01011112222", "이영희", "2025-03-23", "15,000원", "보조배터리"]
                ]
                
                for row_idx, row_data in enumerate(sample_data, 2):
                    for col_idx, cell_value in enumerate(row_data, 1):
                        data_sheet.cell(row=row_idx, column=col_idx).value = cell_value
                
                # 열 너비 조정
                for col in range(1, len(headers) + 1):
                    col_letter = chr(64 + col)
                    data_sheet.column_dimensions[col_letter].width = 15
                
                # 파일 저장
                wb.save(AUTO_TEMPLATE_PATH)
                
            return send_from_directory(DATA_FOLDER, 'automation_template.xlsx', as_attachment=True)
            
        else:
            return jsonify({'success': False, 'message': '유효하지 않은 템플릿 유형입니다.'}), 400
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/parse-recipients-only', methods=['POST'])
def parse_recipients_only():
    """수신자 목록만 파싱하는 API"""
    try:
        # 파일 확인
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '파일이 필요합니다.'}), 400
            
        file = request.files['file']
        if not file or not file.filename:
            return jsonify({'success': False, 'message': '파일이 필요합니다.'}), 400
            
        # 파일 확장자 확인
        file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if file_ext != 'csv':
            return jsonify({'success': False, 'message': 'CSV 형식의 파일만 지원합니다.'}), 400
        
        # 추가: 메시지 내용 가져오기
        text = request.form.get('text', '')
        print(f"parse_recipients_only 요청: text='{text}'")
        
        # 디버깅 모드 활성화
        DEBUG_MODE = os.environ.get('DEBUG_MODE', 'True').lower() == 'true'
        
        # 디버깅 모드일 경우 테스트 응답 반환
        if DEBUG_MODE:
            print("테스트 모드에서 실행 중입니다. 파싱 테스트 응답을 반환합니다.")
            test_recipients = ['01012345678', '01023456789', '01034567890', '01045678901', '01056789012']
            return jsonify({
                'success': True,
                'recipients': test_recipients,
                'count': len(test_recipients),
                'text': text  # 메시지 내용도 반환
            })
            
        # Lambda 요청 데이터 준비
        excel_data = base64.b64encode(file.read()).decode('utf-8')
        
        lambda_data = {
            'type': 'parse_recipients',
            'excel': {
                'data': excel_data,
                'filename': file.filename
            }
        }
        
        # 추가: 메시지 내용이 있으면 Lambda 데이터에 추가
        if text:
            lambda_data['text'] = text
            print(f"Lambda 요청에 text 필드 추가: '{text}'")
        
        # Lambda 함수 호출
        response = requests.post(LAMBDA_FUNCTION_URL, json=lambda_data)
        
        if response.status_code == 200:
            result = response.json()
            return jsonify(result)
        else:
            return jsonify({
                'success': False,
                'message': f'수신자 파싱 실패: {response.text}'
            }), response.status_code
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

def parse_recipients_only_from_file(file, text=''):
    """파일에서 수신자 목록만 파싱하는 함수"""
    try:
        # 파일 확장자 확인
        file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if file_ext != 'csv':
            return {'success': False, 'message': 'CSV 형식의 파일만 지원합니다.'}
        
        # 디버깅 모드 활성화
        DEBUG_MODE = os.environ.get('DEBUG_MODE', 'True').lower() == 'true'
        
        # 디버깅 모드일 경우 테스트 응답 반환
        if DEBUG_MODE:
            print("테스트 모드에서 실행 중입니다. 파싱 테스트 응답을 반환합니다.")
            test_recipients = ['01012345678', '01023456789', '01034567890', '01045678901', '01056789012']
            return {
                'success': True,
                'recipients': test_recipients,
                'count': len(test_recipients),
                'text': text  # 메시지 내용도 반환
            }
            
        # Lambda 요청 데이터 준비
        excel_data = base64.b64encode(file.read()).decode('utf-8')
        
        lambda_data = {
            'type': 'parse_recipients',
            'excel': {
                'data': excel_data,
                'filename': file.filename
            }
        }
        
        # 추가: 메시지 내용이 있으면 Lambda 데이터에 추가
        if text:
            lambda_data['text'] = text
            print(f"Lambda 요청에 text 필드 추가: '{text}'")
        
        # Lambda 함수 호출
        response = requests.post(LAMBDA_FUNCTION_URL, json=lambda_data)
        
        if response.status_code == 200:
            result = response.json()
            return result
        else:
            return {
                'success': False,
                'message': f'수신자 파싱 실패: {response.text}'
            }
            
    except Exception as e:
        return {'success': False, 'message': str(e)}

@app.route('/api/lambda', methods=['POST'])
def lambda_api():
    """람다 API 직접 호출"""
    try:
        print("Lambda API 요청 수신됨")
        
        # 요청 본문 처리 (JSON 또는 폼 데이터)
        if request.is_json:
            print("JSON 요청 데이터 처리")
            data = request.json
            print(f"JSON 데이터: {data}")
        else:
            print("폼 데이터 처리")
            data = {}
            
            # 폼 데이터 처리
            for key in request.form:
                data[key] = request.form[key]
            
            print(f"폼 데이터: {data}")
            
            # 파일 처리
            if 'file' in request.files:
                file = request.files['file']
                # 파일 데이터 처리 (이미지 또는 엑셀 파일)
                if file and file.filename:
                    print(f"파일 업로드 감지: {file.filename}")
                    # 파일 확장자 확인
                    file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                    
                    # 파일 데이터를 base64로 인코딩
                    file_data = base64.b64encode(file.read()).decode('utf-8')
                    
                    # 파일 정보 추가 - excel 필드로 변환
                    data['excel'] = {
                        'filename': file.filename,
                        'data': file_data
                    }
                    print(f"'file' 필드를 'excel' 필드로 변환 완료: {file.filename}")
            
            # 이미지 처리
            if 'image' in request.files:
                image = request.files['image']
                if image and image.filename:
                    print(f"이미지 업로드 감지: {image.filename}")
                    # 이미지 데이터를 base64로 인코딩
                    image_data = base64.b64encode(image.read()).decode('utf-8')
                    
                    # 이미지 정보 추가
                    data['image'] = {
                        'filename': image.filename,
                        'data': image_data
                    }
                    print(f"이미지 데이터 추가 완료: {image.filename}")
        
        # get_template 요청인 경우 로컬에서 직접 처리
        if 'type' in data and data['type'] == 'get_template':
            print("템플릿 다운로드 요청 처리")
            template_path = os.path.join(DATA_FOLDER, 'sample_template.csv')
            
            # 파일 존재 확인
            if not os.path.exists(template_path):
                print(f"오류: 템플릿 파일을 찾을 수 없습니다. (경로: {template_path})")
                return jsonify({
                    'success': False,
                    'message': '템플릿 파일을 찾을 수 없습니다.'
                })
            
            # 파일 읽기
            with open(template_path, 'rb') as f:
                binary_data = f.read()
            
            # Base64 인코딩
            base64_data = base64.b64encode(binary_data).decode('utf-8')
            
            print(f"템플릿 파일 크기: {len(binary_data)} 바이트, Base64 길이: {len(base64_data)}")
            
            return jsonify({
                'success': True,
                'filename': 'sample_template.csv',
                'data': base64_data
            })
        
        # auto_excel_preview 요청 처리 (자동 메시지 템플릿 미리보기)
        elif 'type' in data and data['type'] == 'auto_excel_preview':
            print("자동 메시지 템플릿 미리보기 요청 처리")
            
            # 파일 확인
            if 'file' in request.files:
                file = request.files['file']
                if not file or not file.filename:
                    return jsonify({'success': False, 'message': '파일이 필요합니다.'}), 400
                    
                # 파일 확장자 확인
                if not file.filename.lower().endswith('.xlsx'):
                    return jsonify({'success': False, 'message': 'XLSX 형식의 엑셀 파일만 지원합니다.'}), 400
                
                # 엑셀 데이터 준비
                excel_data = base64.b64encode(file.read()).decode('utf-8')
                lambda_data = {
                    'type': 'auto_excel_preview',
                    'excel': {
                        'data': excel_data,
                        'filename': file.filename
                    }
                }
            # excel 필드가 JSON 문자열로 전달된 경우
            elif 'excel' in data:
                try:
                    if isinstance(data['excel'], str):
                        excel_info = json.loads(data['excel'])
                        lambda_data = {
                            'type': 'auto_excel_preview',
                            'excel': excel_info
                        }
                    else:
                        lambda_data = {
                            'type': 'auto_excel_preview',
                            'excel': data['excel']
                        }
                    print(f"Excel 데이터 형식: {type(lambda_data['excel'])}")
                except Exception as e:
                    print(f"Excel 데이터 파싱 오류: {str(e)}")
                    return jsonify({'success': False, 'message': 'Excel 데이터 형식이 올바르지 않습니다.'}), 400
            else:
                return jsonify({'success': False, 'message': '파일이 필요합니다.'}), 400
            
            # 디버깅 모드 확인
            DEBUG_MODE = os.environ.get('DEBUG_MODE', 'True').lower() == 'true'
            
            # 디버깅 모드일 경우 테스트 응답 반환
            if DEBUG_MODE:
                print("테스트 모드에서 실행 중입니다. 자동 메시지 미리보기 테스트 응답을 반환합니다.")
                preview_data = [
                    {'index': 1, 'phone': '01012345678', 'text': '안녕하세요 홍길동님, 2025-03-22에 주문하신 스마트폰 케이스가 배송되었습니다.'},
                    {'index': 2, 'phone': '01098765432', 'text': '안녕하세요 김철수님, 2025-03-22에 주문하신 블루투스 이어폰이 배송되었습니다.'},
                    {'index': 3, 'phone': '01011112222', 'text': '안녕하세요 이영희님, 2025-03-23에 주문하신 보조배터리가 배송되었습니다.'}
                ]
                return jsonify({
                    'success': True,
                    'total': 3,
                    'preview': preview_data,
                    'message': '자동 메시지 미리보기가 준비되었습니다.'
                })
            
            # Lambda 함수 호출 
            if not LAMBDA_FUNCTION_URL:
                # 로컬 Lambda 함수 직접 호출 대신 process_auto_excel_template 함수 직접 사용
                print("로컬에서 process_auto_excel_template 함수 직접 호출")
                try:
                    # 파일 데이터 처리
                    import lambda_update
                    
                    # 엑셀 데이터 추출
                    if isinstance(lambda_data['excel'], dict) and 'data' in lambda_data['excel']:
                        excel_content = base64.b64decode(lambda_data['excel']['data'])
                        filename = lambda_data['excel'].get('filename', 'uploaded_file.xlsx')
                        
                        print(f"엑셀 파일 크기: {len(excel_content)} bytes, 파일명: {filename}")
                        
                        # 직접 process_auto_excel_template 함수 호출
                        result = lambda_update.process_auto_excel_template(
                            excel_content=excel_content,
                            filename=filename,
                            sender_phone=os.environ.get('SENDER_PHONE', '15881234')
                        )
                        
                        print(f"process_auto_excel_template 결과: {result}")
                        return jsonify(result)
                    else:
                        return jsonify({
                            'success': False,
                            'message': '엑셀 데이터 형식이 올바르지 않습니다.'
                        }), 400
                except Exception as inner_e:
                    print(f"로컬 함수 호출 오류: {str(inner_e)}")
                    traceback.print_exc()
                    return jsonify({
                        'success': False,
                        'message': f'자동 메시지 미리보기 처리 중 오류 발생: {str(inner_e)}'
                    }), 500
            else:
                # Lambda 함수 URL 호출
                print(f"Lambda 함수 URL 호출: {LAMBDA_FUNCTION_URL}")
                print(f"Lambda 요청 데이터: {json.dumps(lambda_data)[:200]}...")
                try:
                    # 타임아웃 증가 및 요청 헤더 추가
                    headers = {'Content-Type': 'application/json'}
                    response = requests.post(
                        LAMBDA_FUNCTION_URL, 
                        json=lambda_data, 
                        headers=headers,
                        timeout=60
                    )
                    
                    print(f"Lambda 응답 상태 코드: {response.status_code}")
                    if response.status_code == 200:
                        result = response.json()
                        return jsonify(result)
                    else:
                        print(f"Lambda 오류 응답: {response.status_code}, {response.text}")
                        return jsonify({
                            'success': False,
                            'message': f'자동 메시지 미리보기 처리 실패: HTTP {response.status_code}, {response.text[:100]}'
                        }), response.status_code
                except requests.exceptions.Timeout:
                    print("Lambda 함수 호출 타임아웃")
                    return jsonify({
                        'success': False,
                        'message': 'Lambda 함수 호출 타임아웃. 네트워크 연결을 확인하세요.'
                    }), 504
                except requests.exceptions.ConnectionError as ce:
                    print(f"Lambda 함수 연결 오류: {str(ce)}")
                    return jsonify({
                        'success': False,
                        'message': f'Lambda 함수 연결 오류: {str(ce)}'
                    }), 502
                except Exception as ex:
                    print(f"Lambda 함수 호출 중 예외 발생: {str(ex)}")
                    traceback.print_exc()
                    return jsonify({
                        'success': False,
                        'message': f'Lambda 함수 호출 중 예외 발생: {str(ex)}'
                    }), 500
            
        # auto_excel_send 요청 처리 (자동 메시지 템플릿 발송)
        elif 'type' in data and data['type'] == 'auto_excel_send':
            print("자동 메시지 템플릿 발송 요청 처리")
            
            # 파일 확인
            if 'file' in request.files:
                file = request.files['file']
                if not file or not file.filename:
                    return jsonify({'success': False, 'message': '파일이 필요합니다.'}), 400
                    
                # 파일 확장자 확인
                if not file.filename.lower().endswith('.xlsx'):
                    return jsonify({'success': False, 'message': 'XLSX 형식의 엑셀 파일만 지원합니다.'}), 400
                
                # 엑셀 데이터 준비
                excel_data = base64.b64encode(file.read()).decode('utf-8')
                lambda_data = {
                    'type': 'auto_excel_send',
                    'excel': {
                        'data': excel_data,
                        'filename': file.filename
                    }
                }
            # excel 필드가 JSON 문자열로 전달된 경우
            elif 'excel' in data:
                try:
                    if isinstance(data['excel'], str):
                        excel_info = json.loads(data['excel'])
                        lambda_data = {
                            'type': 'auto_excel_send',
                            'excel': excel_info
                        }
                    else:
                        lambda_data = {
                            'type': 'auto_excel_send',
                            'excel': data['excel']
                        }
                    print(f"Excel 데이터 형식: {type(lambda_data['excel'])}")
                except Exception as e:
                    print(f"Excel 데이터 파싱 오류: {str(e)}")
                    return jsonify({'success': False, 'message': 'Excel 데이터 형식이 올바르지 않습니다.'}), 400
            else:
                return jsonify({'success': False, 'message': '파일이 필요합니다.'}), 400
            
            try:
                # 디버깅 모드 확인
                DEBUG_MODE = os.environ.get('DEBUG_MODE', 'True').lower() == 'true'
                
                # 디버깅 모드일 경우 테스트 응답 반환
                if DEBUG_MODE:
                    print("테스트 모드에서 실행 중입니다. 자동 메시지 발송 테스트 응답을 반환합니다.")
                    return jsonify({
                        'success': True,
                        'message': '테스트 모드: 메시지가 성공적으로 발송된 것으로 처리됩니다.',
                        'total': 3,
                        'failedCount': 0,
                        'failedList': []
                    })
                
                # 이미지가 있는 경우 처리
                if 'image' in request.files:
                    image = request.files['image']
                    if image and image.filename:
                        image_data = base64.b64encode(image.read()).decode('utf-8')
                        lambda_data['image'] = {
                            'data': image_data,
                            'filename': image.filename
                        }
                # 이미지가 JSON 문자열로 전달된 경우
                elif 'image' in data:
                    try:
                        if isinstance(data['image'], str):
                            image_info = json.loads(data['image'])
                            lambda_data['image'] = image_info
                        else:
                            lambda_data['image'] = data['image']
                    except Exception as e:
                        print(f"이미지 데이터 파싱 오류: {str(e)}")
                
                # Lambda 함수 호출
                if not LAMBDA_FUNCTION_URL:
                    # 로컬 Lambda 함수 직접 호출 대신 process_auto_excel_template 함수 직접 사용
                    print("로컬에서 process_auto_excel_template 함수 직접 호출")
                    try:
                        # 파일 데이터 처리
                        import lambda_update
                        
                        # 엑셀 데이터 추출
                        if isinstance(lambda_data['excel'], dict) and 'data' in lambda_data['excel']:
                            excel_content = base64.b64decode(lambda_data['excel']['data'])
                            filename = lambda_data['excel'].get('filename', 'uploaded_file.xlsx')
                            
                            print(f"엑셀 파일 크기: {len(excel_content)} bytes, 파일명: {filename}")
                            
                            # 직접 process_auto_excel_template 함수 호출
                            result = lambda_update.process_auto_excel_template(
                                excel_content=excel_content,
                                filename=filename,
                                sender_phone=os.environ.get('SENDER_PHONE', '15881234')
                            )
                            
                            print(f"process_auto_excel_template 결과: {result}")
                            return jsonify(result)
                        else:
                            return jsonify({
                                'success': False,
                                'message': '엑셀 데이터 형식이 올바르지 않습니다.'
                            }), 400
                    except Exception as inner_e:
                        print(f"로컬 함수 호출 오류: {str(inner_e)}")
                        traceback.print_exc()
                        return jsonify({
                            'success': False,
                            'message': f'자동 메시지 발송 처리 중 오류 발생: {str(inner_e)}'
                        }), 500
                else:
                    # Lambda 함수 URL 호출
                    print(f"Lambda 함수 URL 호출: {LAMBDA_FUNCTION_URL}")
                    print(f"Lambda 요청 데이터: {json.dumps(lambda_data)[:200]}...")
                    try:
                        # 타임아웃 증가 및 요청 헤더 추가
                        headers = {'Content-Type': 'application/json'}
                        response = requests.post(
                            LAMBDA_FUNCTION_URL, 
                            json=lambda_data, 
                            headers=headers,
                            timeout=60
                        )
                        
                        print(f"Lambda 응답 상태 코드: {response.status_code}")
                        if response.status_code == 200:
                            result = response.json()
                            return jsonify(result)
                        else:
                            print(f"Lambda 오류 응답: {response.status_code}, {response.text}")
                            return jsonify({
                                'success': False,
                                'message': f'자동 메시지 발송 처리 실패: HTTP {response.status_code}, {response.text[:100]}'
                            }), response.status_code
                    except requests.exceptions.Timeout:
                        print("Lambda 함수 호출 타임아웃")
                        return jsonify({
                            'success': False,
                            'message': 'Lambda 함수 호출 타임아웃. 네트워크 연결을 확인하세요.'
                        }), 504
                    except requests.exceptions.ConnectionError as ce:
                        print(f"Lambda 함수 연결 오류: {str(ce)}")
                        return jsonify({
                            'success': False,
                            'message': f'Lambda 함수 연결 오류: {str(ce)}'
                        }), 502
                    except Exception as ex:
                        print(f"Lambda 함수 호출 중 예외 발생: {str(ex)}")
                        traceback.print_exc()
                        return jsonify({
                            'success': False,
                            'message': f'Lambda 함수 호출 중 예외 발생: {str(ex)}'
                        }), 500
            except Exception as e:
                print(f"자동 메시지 발송 처리 중 오류 발생: {str(e)}")
                import traceback
                traceback.print_exc()
                return jsonify({
                    'success': False,
                    'message': f'자동 메시지 발송 처리 중 오류 발생: {str(e)}'
                }), 500
            
        # 나머지 요청은 Lambda 함수로 전달
        DEBUG_MODE = os.environ.get('DEBUG_MODE', 'False').lower() == 'true'
        
        if DEBUG_MODE:
            # 디버그 모드일 경우 로컬 람다 함수 호출
            print("DEBUG_MODE: 로컬에서 lambda_function 호출")
            import lambda_update as lambda_function
            lambda_response = lambda_function.lambda_handler({
                'body': json.dumps(data) if isinstance(data, dict) else data
            }, {})
            
            print(f"로컬 Lambda 응답: {lambda_response}")
            return jsonify(lambda_response)
        else:
            # 프로덕션 모드일 경우 Lambda 함수 URL 호출
            print(f"프로덕션 모드: Lambda 함수 URL 호출: {LAMBDA_FUNCTION_URL}")
            response = requests.post(LAMBDA_FUNCTION_URL, json=data, timeout=30)  # 타임아웃 30초로 설정
            
            if response.status_code == 200:
                return jsonify(response.json())
            else:
                print(f"Lambda 오류 응답: {response.status_code}, {response.text}")
                return jsonify({
                    'success': False,
                    'message': f'Lambda 함수 호출 실패: {response.text}'
                }), response.status_code
    
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'오류 발생: {str(e)}'
        }), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True) 