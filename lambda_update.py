import json
import base64
import os
import urllib.request
import urllib.error
import urllib.parse
import hmac
import hashlib
import uuid
import datetime
import io
import mimetypes
import platform
import sys
import requests
import csv
import re
import pandas as pd
import traceback
from io import StringIO
from datetime import datetime, timezone, timedelta

# 변경 이력
# -----------------------------------
# 2023-07-01: 초기 구현
# 2023-10-15: 자동 메시지 기능 추가
# 2024-05-13: auto_excel_preview 타입으로 통일, bulk_excel 타입 참조 제거
# -----------------------------------

# 솔라피 API URL 상수 추가
API_BASE_URL = "https://api.solapi.com/messages/v4"
FILE_UPLOAD_URL = "https://api.solapi.com/storage/v1/files"

def get_auth_header(api_key, api_secret):
    """HTTP 요청 인증을 위한, HMAC 서명 기반 헤더를 생성합니다."""
    date = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%fZ')
    salt = str(uuid.uuid4())
    
    # HMAC 서명 생성
    signature_message = date + salt
    signature = hmac.new(
        api_secret.encode(),
        signature_message.encode(),
        hashlib.sha256
    ).hexdigest()
    
    # 헤더 생성
    return {
        'Authorization': f'HMAC-SHA256 apiKey={api_key}, date={date}, salt={salt}, signature={signature}',
        'Content-Type': 'application/json'
    }

def upload_file(api_key, api_secret, file_content, filename):
    """솔라피 API에 파일을 업로드합니다."""
    print(f"파일 업로드 시작: filename={filename}, 크기={len(file_content)} bytes")
    
    try:
        # 파일 크기 제한 (MMS 이미지는 200KB 이하)
        if len(file_content) > 200 * 1024:  # 200KB를 바이트로 환산
            print(f"파일 크기 초과: {len(file_content)} bytes (최대 200KB)")
            return None, "FILE_TOO_LARGE"
        
        # 파일 형식 확인
        content_type, _ = mimetypes.guess_type(filename)
        print(f"파일 MIME 타입: {content_type}")
        
        # JPG 파일만 허용
        if not content_type or content_type not in ['image/jpeg', 'image/jpg']:
            print(f"지원되지 않는 파일 형식: {content_type}")
            print("JPG 형식의 파일만 지원됩니다.")
            return None, "INVALID_FILE_TYPE"
        
        # base64 인코딩
        file_base64 = base64.b64encode(file_content).decode('utf-8')
        
        # Authorization 헤더 생성
        headers = get_auth_header(api_key, api_secret)
        
        # 요청 페이로드 구성
        payload = {
            'file': file_base64,
            'type': 'MMS',
            'name': filename  # 원본 파일명 그대로 사용
        }
        
        print(f"요청 URL: {FILE_UPLOAD_URL}")
        print(f"요청 헤더: Content-Type={headers['Content-Type']}")
        print(f"페이로드 키: {list(payload.keys())}")
        print(f"파일명: {filename}")
        
        # 요청 전송
        response = requests.post(
            FILE_UPLOAD_URL,
            headers=headers,
            json=payload
        )
        
        print(f"API 응답: status_code={response.status_code}, text={response.text}")
        
        if response.status_code == 200:
            result = response.json()
            if "fileId" in result:
                file_id = result["fileId"]
                print(f"파일 업로드 성공: fileId={file_id}")
                return file_id, ""
            else:
                print("fileId가 응답에 없음")
                return None, "NO_FILE_ID"
        else:
            print(f"API 오류 응답: {response.status_code}")
            return None, response.text
    except Exception as e:
        print(f"파일 업로드 중 예외 발생: {str(e)}")
        return None, str(e)

def send_mms(api_key, api_secret, to, from_number, text, image_id):
    """이미지가 첨부된 MMS를 발송합니다."""
    # 인증 헤더 생성
    headers = get_auth_header(api_key, api_secret)
    
    # 이미지 ID 확인
    if not image_id:
        print("이미지 ID가 없습니다. MMS를 발송할 수 없습니다.")
        return {
            "success": False,
            "message": "이미지 ID가 없어 MMS를 발송할 수 없습니다."
        }
    
    # 요청 데이터 구성
    data = {
        "message": {
            "to": to,
            "from": from_number,
            "text": text,
            "type": "MMS",  # 명시적으로 MMS 타입 설정
            "imageId": image_id
        }
    }
    
    print(f"MMS 요청 데이터: {json.dumps(data)}")
    
    # API 요청
    api_url = "https://api.solapi.com/messages/v4/send"
    try:
        # requests를 사용한 요청
        response = requests.post(api_url, headers=headers, json=data)
        print(f"Solapi MMS 응답: status_code={response.status_code}, text={response.text}")
        
        if response.status_code == 200:
            result = response.json()
            return {
                "success": True, 
                "message": "MMS가 성공적으로 발송되었습니다.",
                "result": result
            }
        else:
            return {
                "success": False, 
                "message": f"MMS 발송 실패: {response.text}"
            }
    except Exception as e:
        print(f"MMS 발송 중 오류 발생: {str(e)}")
        return {
            "success": False, 
            "message": f"MMS 발송 중 오류: {str(e)}"
        }

def send_sms(api_key, api_secret, to, from_number, text):
    """단일 SMS를 발송합니다."""
    # 인증 헤더 생성
    headers = get_auth_header(api_key, api_secret)
    
    # 메시지 타입 결정 (SMS 또는 LMS)
    message_type = "SMS"
    if text and len(text) > 90:
        message_type = "LMS"
    
    print(f"메시지 타입: {message_type}")
    
    # 요청 데이터 구성
    data = {
        "message": {
            "to": to,
            "from": from_number,
            "text": text,
            "type": message_type
        }
    }
    
    # API 요청
    api_url = "https://api.solapi.com/messages/v4/send"
    try:
        # requests를 사용한 요청
        response = requests.post(api_url, headers=headers, json=data)
        print(f"Solapi 응답: status_code={response.status_code}, text={response.text[:200]}")
        
        if response.status_code == 200:
            result = response.json()
            return {
                "success": True, 
                "message": "메시지가 성공적으로 발송되었습니다.",
                "result": result
            }
        else:
            return {
                "success": False, 
                "message": f"발송 실패: {response.text}"
            }
    except Exception as e:
        print(f"SMS 발송 중 오류 발생: {str(e)}")
        return {
            "success": False, 
            "message": f"발송 중 오류: {str(e)}"
        }

def send_many_messages(api_key, api_secret, messages):
    """다수의 메시지를 발송합니다."""
    if not messages:
        return {"error": "메시지가 없습니다."}
        
    # 인증 헤더 생성
    headers = get_auth_header(api_key, api_secret)
    
    # 요청 데이터 구성
    data = {"messages": messages}
    
    print(f"대량 메시지 발송 요청: {json.dumps(data)}")
    
    # API 요청
    api_url = "https://api.solapi.com/messages/v4/send-many"
    try:
        # requests를 사용한 요청
        response = requests.post(api_url, headers=headers, json=data)
        print(f"Solapi 대량 발송 응답: status_code={response.status_code}, text={response.text[:200]}")
        
        if response.status_code == 200:
            result = response.json()
            return result
        else:
            return {"error": response.text}
    except Exception as e:
        print(f"대량 메시지 발송 중 오류 발생: {str(e)}")
        return {"error": str(e)}

def parse_recipients_only(excel_data, filename=None):
    """CSV 파일에서 수신자 번호만 추출합니다."""
    try:
        # CSV 파일 읽기
        csv_text = excel_data.decode('utf-8-sig')  # BOM 처리
        csv_file = StringIO(csv_text)
        csv_reader = csv.reader(csv_file)
        
        print(f"CSV 파일 읽기 시작: {filename}")
        
        # 헤더 읽기
        headers = next(csv_reader, None)
        if not headers:
            print("CSV 파일에 헤더가 없습니다.")
            return {"success": False, "message": "CSV 파일에 헤더가 없습니다."}
            
        print(f"CSV 헤더: {headers}")
        
        # B열 (인덱스 1)에서 전화번호 추출
        recipients = []
        names = []  # 이름 목록도 추출
        for row_idx, row in enumerate(csv_reader, 1):
            try:
                if len(row) > 1:  # B열이 있는지 확인
                    phone = str(row[1]).strip()  # B열 (인덱스 1)
                    # 숫자만 추출
                    phone = ''.join(c for c in phone if c.isdigit())
                    
                    # 이름 추출 (A열)
                    name = row[0].strip() if len(row) > 0 else ""
                    
                    if phone and len(phone) >= 10:
                        recipients.append(phone)
                        names.append(name)
                        print(f"행 {row_idx}: 전화번호 추출 성공 - {phone}, 이름: {name}")
                    else:
                        print(f"행 {row_idx}: 유효하지 않은 전화번호 형식 - {row[1]}")
                else:
                    print(f"행 {row_idx}: B열이 없음")
            except Exception as row_error:
                print(f"행 {row_idx} 처리 중 오류: {str(row_error)}")
                continue
        
        print(f"CSV 처리 완료: 총 {len(recipients)}개의 전화번호 추출")
        
        if not recipients:
            return {
                "success": False,
                "message": "유효한 전화번호를 찾을 수 없습니다."
            }
        
        return {
            "success": True,
            "recipients": recipients,
            "names": names,
            "count": len(recipients)
        }
    except Exception as e:
        print(f"CSV 파일 처리 중 오류 발생: {str(e)}")
        return {"success": False, "message": str(e)}

def lambda_handler(event, context):
    try:
        # 환경 변수 가져오기
        api_key = os.environ.get('API_KEY', '')
        api_secret = os.environ.get('API_SECRET', '')
        sender_phone = os.environ.get('SENDER_PHONE', '')
        aws_access_key = os.environ.get('MY_AWS_ACCESS_KEY', '')
        aws_secret_key = os.environ.get('MY_AWS_SECRET_KEY', '')
        aws_bucket_name = os.environ.get('MY_AWS_BUCKET_NAME', '')
        aws_region = os.environ.get('MY_AWS_REGION', 'ap-northeast-2')
        
        # 디버깅: 환경 변수 로깅 (API 키와 시크릿은 보안상 일부만 표시)
        print(f"API_KEY: {api_key[:4]}...{api_key[-4:] if len(api_key) > 8 else ''}")
        print(f"API_SECRET: {api_secret[:4]}...{api_secret[-4:] if len(api_secret) > 8 else ''}")
        print(f"SENDER_PHONE: {sender_phone}")
        print(f"MY_AWS_ACCESS_KEY: {aws_access_key[:4]}...{aws_access_key[-4:] if len(aws_access_key) > 8 else ''}")
        print(f"MY_AWS_SECRET_KEY: {aws_secret_key[:4]}...{aws_secret_key[-4:] if len(aws_secret_key) > 8 else ''}")
        print(f"MY_AWS_BUCKET_NAME: {aws_bucket_name}")
        print(f"MY_AWS_REGION: {aws_region}")
        
        # 디버깅: 입력 이벤트 로깅
        print(f"받은 이벤트: {json.dumps(event)}")
        
        # Lambda URL을 통한 요청 처리
        if 'body' in event:
            # Lambda URL을 통한 요청
            print("Lambda URL 요청 감지됨")
            try:
                if isinstance(event['body'], str):
                    # Lambda URL 요청이면 JSON 문자열로 받음
                    print("문자열 형태의 본문 처리")
                    body = json.loads(event['body'])
                else:
                    # API Gateway 연동 시에는 이미 파싱된 객체로 받을 수 있음
                    print("객체 형태의 본문 처리")
                    body = event['body']
            except Exception as e:
                print(f"JSON 파싱 오류: {str(e)}, body: {event['body']}")
                return {
                    'statusCode': 400,
                    'body': json.dumps({
                        'success': False,
                        'message': f'잘못된 요청 형식입니다: {str(e)}'
                    }),
                }
        else:
            # 직접 Lambda 호출
            print("직접 Lambda 호출 감지됨")
            body = event

        print(f"처리할 본문 데이터: {json.dumps(body)}")

        # 요청 타입 확인
        if 'type' not in body:
            return {
                'success': False,
                'message': '요청 타입이 누락되었습니다.'
            }

        request_type = body['type']
        print(f"요청 타입: '{request_type}'")

        # 디버깅용 ping 요청 처리
        if request_type == 'ping':
            return {
                'success': True,
                'message': 'Lambda 함수가 정상적으로 응답했습니다.',
                'request': body
            }
            
        # 자동메시지 엑셀 미리보기 요청 처리
        elif request_type == 'auto_excel_preview':
            print("자동메시지 엑셀 미리보기 요청 처리 시작")
            
            if 'excel' not in body or not body['excel']:
                return {
                    'success': False,
                    'message': '엑셀 파일이 필요합니다.'
                }
                
            excel_data = body['excel']
            if isinstance(excel_data, str):
                excel_data = json.loads(excel_data)
                
            if 'data' not in excel_data or not excel_data['data']:
                return {
                    'success': False,
                    'message': '엑셀 파일 데이터가 비어있습니다.'
                }
                
            excel_content = base64.b64decode(excel_data['data'])
            excel_filename = excel_data.get('filename', 'excel.xlsx')
            
            # 자동 메시지 템플릿 처리
            result = process_auto_excel_template(excel_content, excel_filename, body, sender_phone)
            print(f"자동메시지 엑셀 미리보기 결과: {json.dumps(result)}")
            
            return result
            
        elif request_type == 'auto_excel_send':
            print("자동메시지 엑셀 발송 요청 처리 시작")
            
            if 'excel' not in body or not body['excel']:
                return {
                    'success': False,
                    'message': '엑셀 파일이 필요합니다.'
                }
                
            excel_data = body['excel']
            excel_content = base64.b64decode(excel_data['data'])
            excel_filename = excel_data['filename']
            
            # 자동 메시지 템플릿 처리
            template_result = process_auto_excel_template(excel_content, excel_filename, body, sender_phone)
            
            if not template_result['success']:
                return template_result
                
            recipients = template_result['recipients']
            
            if not recipients:
                return {
                    'success': False,
                    'message': '발송할 메시지가 없습니다.'
                }
            
            # 이미지가 있는 경우 처리
            if 'image' in body and body['image']:
                image_data = body['image']
                # 문자열로 전달된 경우 JSON으로 파싱
                if isinstance(image_data, str):
                    try:
                        image_data = json.loads(image_data)
                    except Exception as e:
                        print(f"이미지 데이터 파싱 오류: {str(e)}")
                        return {
                            'success': False,
                            'message': f'이미지 데이터 형식 오류: {str(e)}'
                        }
                
                if isinstance(image_data, dict) and 'data' in image_data and image_data['data']:
                    image_content = base64.b64decode(image_data['data'])
                    image_filename = image_data.get('filename', 'image.jpg')
                    
                    # 솔라피 API에 이미지 업로드
                    image_id, error = upload_file(
                        api_key, 
                        api_secret, 
                        file_content=image_content, 
                        filename=image_filename
                    )
                    
                    if image_id:
                        # 모든 메시지에 이미지 추가
                        for recipient in recipients:
                            recipient['imageId'] = image_id
                            recipient['type'] = 'MMS'
                    else:
                        return {
                            'success': False,
                            'message': f'이미지 업로드 실패: {error}'
                        }
            
            # 메시지 발송
            for recipient in recipients:
                if 'text' in recipient and recipient['text']:
                    recipient['text'] = format_message_for_sms(recipient['text'])
            
            result = send_many_messages(
                os.environ.get('API_KEY', ''),
                os.environ.get('API_SECRET', ''),
                recipients
            )
            
            # 응답 결과 가공
            response = {
                'success': True,
                'total': len(recipients),
                'failedCount': 0,
                'failedList': [],
                'message': '자동 메시지가 성공적으로 발송되었습니다.'
            }
            
            # 실패 메시지 처리
            if "failedMessageList" in result and result["failedMessageList"]:
                failed_list = result["failedMessageList"]
                response["failedCount"] = len(failed_list)
                
                error_messages = {
                    "ValidationError": "유효성 검사 오류",
                    "HttpError": "HTTP 오류",
                    "InsufficientBalance": "잔액 부족",
                    "NotEnoughBalance": "잔액 부족",
                    "RateLimitError": "요청 한도 초과",
                    "ServerError": "서버 오류",
                    "InvalidPhoneNumber": "유효하지 않은 전화번호",
                    "InvalidFrom": "발신번호 오류",
                    "BlockedNumber": "차단된 번호"
                }
                
                for failed in failed_list:
                    error_info = {
                        "to": failed.get("to", "알 수 없음"),
                        "reason": "알 수 없는 오류"
                    }
                    
                    if "errorCode" in failed:
                        code = failed["errorCode"]
                        if code in error_messages:
                            error_info["reason"] = error_messages[code]
                        else:
                            error_info["reason"] = f"오류 코드: {code}"
                    
                    response["failedList"].append(error_info)
            
            return response
        
        elif request_type == 'test':
            return {
                'success': True,
                'message': 'Lambda 함수 테스트 성공'
            }
        
        elif request_type == 'single':
            # 단일 메시지 처리
            to = body.get('to', '')
            message = body.get('message', '')
            
            if not to or not message:
                return {
                    'success': False,
                    'message': '수신번호와 메시지 내용이 필요합니다.'
                }
            
            # API 키 확인
            if not api_key or not api_secret or not sender_phone:
                return {
                    'success': False,
                    'message': 'Solapi API 키, 시크릿 또는 발신번호가 설정되지 않았습니다.'
                }
            
            # 이미지 처리
            image_id = None
            if 'image' in body:
                print(f"이미지 필드 발견: {json.dumps(body['image'])}")
                
                image_data = body['image']
                # 이미지 데이터가 비어있거나 null인 경우 처리
                if image_data is None or (isinstance(image_data, dict) and len(image_data) == 0):
                    print("이미지 필드가 비어있습니다. 일반 SMS로 발송합니다.")
                elif not isinstance(image_data, dict):
                    print(f"이미지 데이터 형식 오류: {type(image_data)}")
                    return {
                        'success': False,
                        'message': f'이미지 데이터 형식 오류: 딕셔너리 형태여야 합니다.'
                    }
                elif 'data' not in image_data or not image_data.get('data'):
                    print("이미지 데이터에 'data' 필드가 없거나 비어있습니다.")
                    return {
                        'success': False,
                        'message': '이미지 데이터에는 base64로 인코딩된 data 필드가 필요합니다.'
                    }
                else:
                    try:
                        print(f"이미지 데이터 처리 시작: data 길이={len(image_data['data'])}")
                        image_content = base64.b64decode(image_data['data'])
                        image_filename = image_data.get('filename', 'image.jpg')
                        
                        print(f"이미지 디코딩 완료: 크기={len(image_content)} bytes, 파일명={image_filename}")
                        
                        # 솔라피 API에 이미지 업로드
                        image_id, error = upload_file(
                            api_key, 
                            api_secret, 
                            image_content, 
                            image_filename
                        )
                        
                        if not image_id:
                            return {
                                'success': False,
                                'message': f'이미지 업로드 실패: {error}'
                            }
                        
                        # MMS 발송
                        result = send_mms(api_key, api_secret, to, sender_phone, message, image_id)
                        print(f"MMS 발송 결과: {json.dumps(result)}")
                        return result
                    except Exception as e:
                        print(f"MMS 처리 중 오류 발생: {str(e)}")
                        return {
                            'success': False,
                            'message': f'MMS 처리 중 오류: {str(e)}'
                        }
            
            # 일반 SMS/LMS 발송
            result = send_sms(api_key, api_secret, to, sender_phone, message)
            print(f"SMS 발송 결과: {json.dumps(result)}")
            return result
            
        elif request_type == 'parse_recipients':
            if 'excel' not in body or not body['excel']:
                return {
                    'success': False,
                    'message': '파일이 필요합니다.'
                }
                    
            try:
                excel_data = body['excel']
                csv_content = base64.b64decode(excel_data['data'])
                csv_filename = excel_data.get('filename', 'recipients.csv')
                
                print(f"CSV 파일 읽기 시작: {csv_filename}")
                
                # 전체 body 디버깅 출력 (text 필드 확인용)
                print("=" * 50)
                print(f"처리할 본문 데이터 타입: {type(body)}")
                print(f"처리할 본문 데이터 키: {list(body.keys())}")
                
                # 사용자 메시지 텍스트 확인 - 여러 가능한 위치 확인
                user_text = None
                
                # 1. 최상위 text 필드
                if 'text' in body:
                    user_text = body['text']
                    print(f"✅ 최상위 text 필드 발견: '{user_text}'")
                
                # 2. form 필드 내부
                elif 'form' in body and isinstance(body['form'], dict) and 'text' in body['form']:
                    user_text = body['form']['text']
                    print(f"✅ form.text 필드 발견: '{user_text}'")
                
                # 3. body 필드 내부
                elif 'body' in body and isinstance(body['body'], dict) and 'text' in body['body']:
                    user_text = body['body']['text']
                    print(f"✅ body.text 필드 발견: '{user_text}'")
                
                # 4. request 필드 내부
                elif 'request' in body and isinstance(body['request'], dict):
                    request_data = body['request']
                    if 'text' in request_data:
                        user_text = request_data['text']
                        print(f"✅ request.text 필드 발견: '{user_text}'")
                
                # 기본 메시지 설정
                if not user_text:
                    user_text = "안녕하세요, [솔라피] 문자 발송 서비스 메시지입니다."
                    print(f"⚠️ 요청에 text 필드가 없습니다. 기본 메시지가 사용될 수 있습니다.")
                
                # 수신자 목록 추출
                result = parse_recipients_only(csv_content, csv_filename)
                if not result['success']:
                    return result
                
                recipients = result['recipients']
                names = result.get('names', [])
                print(f"추출된 수신자 수: {len(recipients)}")
                
                # 미리보기 정보 구성
                preview_messages = []
                for i, recipient in enumerate(recipients[:5]):  # 처음 5개만 미리보기
                    name = names[i] if i < len(names) else ""
                    
                    # 메시지 구성
                    message_text = user_text
                    
                    preview = {
                        'index': i + 1,
                        'name': name,
                        'phone': recipient,
                        'text': message_text,
                        'type': 'LMS' if len(message_text) > 90 else 'SMS'
                    }
                    preview_messages.append(preview)
                
                # 응답 구성 - text 필드 포함
                return {
                    'success': True,
                    'recipients': recipients,
                    'count': len(recipients),
                    'preview': preview_messages,
                    'message': 'CSV 파일에서 수신자 정보를 성공적으로 읽었습니다.',
                    'text': user_text  # 사용자 메시지 반환
                }
                
            except Exception as e:
                print(f"CSV 파일 처리 중 오류 발생: {str(e)}")
                return {
                    'success': False,
                    'message': f'CSV 파일 처리 중 오류 발생: {str(e)}'
                }
                
        elif request_type == 'send_message':
            # CSV에서 추출한 번호들에게 일괄 메시지 발송
            try:
                # 필수 필드 확인
                if 'text' not in body or not body['text']:
                    return {
                        'success': False,
                        'message': '메시지 내용이 필요합니다.'
                    }
                
                # 수신자 목록 가져오기 (여러 가능한 소스에서 확인)
                recipients = []
                
                # 1. 직접 전달된 recipients 배열이 있는 경우
                if 'recipients' in body and body['recipients']:
                    try:
                        # recipients가 문자열인 경우 JSON으로 파싱
                        if isinstance(body['recipients'], str):
                            recipients = json.loads(body['recipients'])
                        else:
                            recipients = body['recipients']
                        print(f"요청에서 직접 수신자 목록 사용: {len(recipients)}명")
                    except Exception as parse_error:
                        print(f"수신자 목록 파싱 오류: {str(parse_error)}")
                        print(f"원본 수신자 데이터 타입: {type(body['recipients'])}")
                        print(f"원본 수신자 데이터: {body['recipients'][:100] if isinstance(body['recipients'], str) else str(body['recipients'])}")
                        return {
                            'success': False,
                            'message': f'수신자 목록 형식 오류: {str(parse_error)}'
                        }
                
                # 2. excel 데이터가 본문에 포함된 경우 (Lambda API Gateway 호출)
                elif 'excel' in body and isinstance(body['excel'], dict) and 'data' in body['excel']:
                    # CSV 파일에서 수신자 목록 추출
                    excel_data = body['excel']
                    csv_content = base64.b64decode(excel_data['data'])
                    csv_filename = excel_data.get('filename', 'recipients.csv')
                    
                    print(f"CSV 파일에서 수신자 목록 추출: {csv_filename}")
                    
                    # CSV 파일 처리하여 수신자 목록 추출
                    recipients_result = parse_recipients_only(csv_content, csv_filename)
                    if recipients_result['success']:
                        recipients = recipients_result['recipients']
                        print(f"CSV 파일에서 추출한 수신자: {len(recipients)}명")
                    else:
                        print(f"CSV 파일 처리 실패: {recipients_result['message']}")
                
                # 3. file 필드가 전달된 경우 (이 경우는 app.py에서 처리함)
                elif 'file' in body and isinstance(body['file'], dict) and 'data' in body['file']:
                    # 파일에서 수신자 목록 추출
                    file_data = body['file']
                    file_content = base64.b64decode(file_data['data'])
                    file_name = file_data.get('filename', 'recipients.csv')
                    
                    print(f"file 필드에서 수신자 목록 추출: {file_name}")
                    
                    recipients_result = parse_recipients_only(file_content, file_name)
                    if recipients_result['success']:
                        recipients = recipients_result['recipients']
                        print(f"file에서 추출한 수신자: {len(recipients)}명")
                    else:
                        print(f"file 처리 실패: {recipients_result['message']}")
                
                # 4. 대안으로 parse_recipients 이벤트 처리 결과에서 추출한 정보 사용
                elif 'parse_recipients_result' in body and isinstance(body['parse_recipients_result'], dict):
                    result = body['parse_recipients_result']
                    if 'recipients' in result and result['recipients']:
                        recipients = result['recipients']
                        print(f"parse_recipients_result에서 수신자 목록 사용: {len(recipients)}명")
                
                # 5. recipients[0], recipients[1] 형태의 키가 있는 경우 처리
                else:
                    # recipients[0], recipients[1] 같은 형태의 키 찾기
                    recipient_keys = [k for k in body.keys() if k.startswith('recipients[') and k.endswith(']')]
                    if recipient_keys:
                        print(f"recipients[n] 형태의 키 발견: {len(recipient_keys)}개")
                        # 키를 정렬하여 순서대로 처리
                        recipient_keys.sort(key=lambda k: int(k.replace('recipients[', '').replace(']', '')))
                        for key in recipient_keys:
                            if body[key]:  # 빈 값이 아닌 경우만 추가
                                recipients.append(body[key])
                        print(f"recipients[n] 형태에서 추출한 수신자: {len(recipients)}명")
                
                if not recipients:
                    print("유효한 수신자 정보를 찾을 수 없습니다.")
                    print(f"요청 본문 키: {list(body.keys())}")
                    
                    # excel 키 내용 확인
                    if 'excel' in body:
                        print(f"excel 키 타입: {type(body['excel'])}")
                        if isinstance(body['excel'], dict):
                            print(f"excel 키 내용: {list(body['excel'].keys())}")
                    
                    return {
                        'success': False,
                        'message': '유효한 수신자 정보가 없습니다. 수신자 목록 또는 CSV 파일이 필요합니다.'
                    }
                
                # 텍스트 메시지 가져오기
                text = body['text']
                print(f"메시지 발송 요청: {len(recipients)}명, 내용: '{text}'")
                
                # 이미지 처리
                image_id = None
                if 'image' in body and body['image']:
                    image_data = body['image']
                    # 문자열로 전달된 경우 JSON으로 파싱
                    if isinstance(image_data, str):
                        try:
                            image_data = json.loads(image_data)
                        except Exception as e:
                            print(f"이미지 데이터 파싱 오류: {str(e)}")
                            return {
                                'success': False,
                                'message': f'이미지 데이터 형식 오류: {str(e)}'
                            }
                    
                    if isinstance(image_data, dict) and 'data' in image_data and image_data['data']:
                        image_content = base64.b64decode(image_data['data'])
                        image_filename = image_data.get('filename', 'image.jpg')
                        
                        # 솔라피 API에 이미지 업로드
                        image_id, error = upload_file(
                            api_key, 
                            api_secret, 
                            file_content=image_content, 
                            filename=image_filename
                        )
                        
                        if not image_id:
                            return {
                                'success': False,
                                'message': f'이미지 업로드 실패: {error}'
                            }
                
                # 메시지 객체 생성
                messages = []
                for recipient in recipients:
                    # 메시지 타입 자동 감지
                    msg_type = 'MMS' if image_id else ('LMS' if len(text) > 90 else 'SMS')
                    
                    # 메시지 객체 생성
                    phone_clean = recipient.replace('-', '')
                    # 국제번호 형식으로 시작하는 경우 국내 형식으로 변환
                    if phone_clean.startswith('82') and len(phone_clean) >= 10:
                        phone_clean = '0' + phone_clean[2:]  # 82 제거하고 앞에 0 추가
                        print(f"국제번호 형식 -> 국내번호 형식 변환: {phone_clean}")
                    
                    message = {
                        'to': phone_clean,  # 정제된 번호 사용
                        'from': sender_phone.replace('-', ''),  # 하이픈 제거
                        'text': format_message_for_sms(text),  # 메시지 포맷팅 적용
                        'type': msg_type  # 명시적으로 메시지 타입 지정
                    }
                    
                    # subject는 SMS에서 사용할 수 없음 - LMS와 MMS 타입에만 추가
                    if msg_type in ['LMS', 'MMS']:
                        message['subject'] = "[자동메시지]"
                    
                    if image_id:
                        message['imageId'] = image_id
                        
                    messages.append(message)
                
                print(f"발송할 메시지 수: {len(messages)}")
                print(f"메시지 내용: '{text}'")
                
                # 대량 메시지 발송 API 호출
                result = send_many_messages(api_key, api_secret, messages)
                
                # 응답 결과 가공
                response = {
                    'success': True,
                    'total': len(messages),
                    'message': '대량 메시지가 성공적으로 발송되었습니다.',
                    'text': text
                }
                
                # 실패 메시지 처리
                if isinstance(result, dict) and "failedMessageList" in result and result["failedMessageList"]:
                    failed_list = result["failedMessageList"]
                    response["failedCount"] = len(failed_list)
                    
                    error_messages = {
                        "ValidationError": "유효성 검사 오류",
                        "HttpError": "HTTP 오류",
                        "InsufficientBalance": "잔액 부족",
                        "NotEnoughBalance": "잔액 부족",
                        "RateLimitError": "요청 한도 초과",
                        "ServerError": "서버 오류",
                        "InvalidPhoneNumber": "유효하지 않은 전화번호",
                        "InvalidFrom": "발신번호 오류",
                        "BlockedNumber": "차단된 번호"
                    }
                    
                    failed_list_info = []
                    for failed in failed_list:
                        error_info = {
                            "to": failed.get("to", "알 수 없음"),
                            "reason": "알 수 없는 오류"
                        }
                        
                        if "errorCode" in failed:
                            code = failed["errorCode"]
                            if code in error_messages:
                                error_info["reason"] = error_messages[code]
                            else:
                                error_info["reason"] = f"오류 코드: {code}"
                        
                        failed_list_info.append(error_info)
                
                return response
                
            except Exception as e:
                print(f"메시지 발송 중 오류 발생: {str(e)}")
                return {
                    'success': False,
                    'message': f'메시지 발송 중 오류 발생: {str(e)}'
                }
        
        elif request_type == 'get_template':
            template_path = os.path.join('data', 'sample_template.csv')
            
            # Local development path
            if os.path.exists(template_path):
                binary_data = None
                with open(template_path, 'rb') as f:
                    binary_data = f.read()
            else:
                base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                template_path = os.path.join(base_dir, 'solapi_project', 'data', 'sample_template.csv')
                # Lambda environment path
                if not os.path.exists(template_path):
                    template_path = '/tmp/sample_template.csv'
                
                binary_data = None
                with open(template_path, 'rb') as f:
                    binary_data = f.read()
                    
            if binary_data:
                base64_data = base64.b64encode(binary_data).decode('utf-8')
                return {
                    'success': True,
                    'filename': 'sample_template.csv',
                    'data': base64_data
                }
        
        else:
            return {
                'success': False,
                'message': f'알 수 없는 요청 타입: {request_type}'
            }
            
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        return {
            'success': False,
            'message': f'오류 발생: {str(e)}'
        }

# 메시지 포맷팅 함수 추가
def format_message_for_sms(text):
    """
    문자 메시지에 최적화된 포맷팅을 적용합니다.
    솔라피 메시지 전송 시 줄바꿈과 공백이 유지되도록 처리합니다.
    """
    if not text:
        return text
    
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    text = text.replace('\xa0', ' ')
    text = text.replace(': ', ':').replace(':', ': ')
    while "  " in text: text = text.replace("  ", " ")
    text = text.replace('\n', '\r\n')
    print(f"포맷팅 후 메시지 확인: {repr(text)}")
    return text

# 자동메시지 처리를 위한 함수 추가
def process_auto_excel_template(excel_content, filename=None, body=None, sender_phone=None):
    """자동 메시지 템플릿을 처리하여 메시지를 생성합니다."""
    try:
        print("=" * 80)
        print("자동메시지 엑셀 처리 시작")
        print(f"파일명: {filename}, 크기: {len(excel_content)} bytes")
        print("=" * 80)

        if not excel_content or len(excel_content) == 0:
            print("엑셀 데이터가 비어 있습니다.")
            return {
                'success': False,
                'message': '엑셀 데이터가 비어 있습니다.'
            }
            
        # 엑셀 파일 임시 저장
        temp_excel_path = '/tmp/temp_excel.xlsx'
        with open(temp_excel_path, 'wb') as f:
            f.write(excel_content)
        print(f"엑셀 파일 임시 저장 완료: {temp_excel_path}")
        
        # 샘플 템플릿 확인 (A2 셀)
        has_template_from_a2 = False
        sample_template = "안녕하세요 {{이름}}님, {{주문일자}}에 주문하신 상품의 금액은 {{주문금액}}원입니다."
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(temp_excel_path, read_only=True, data_only=True)
            
            # 시트 이름 출력
            print(f"엑셀 파일에 있는 모든 시트: {wb.sheetnames}")
            
            # 자동메시지 시트 찾기 (여러 가능한 이름 시도)
            template_sheet = None
            for sheet_name in ['자동메시지', '자동메시지템플릿', 'sample', 'template', 'Sheet1']:
                if sheet_name in wb.sheetnames:
                    template_sheet = wb[sheet_name]
                    print(f"템플릿 시트 '{sheet_name}' 발견됨")
                    break
            
            if template_sheet:
                a2_value = template_sheet['A2'].value

                if a2_value and a2_value.strip():
                    print(f"원본 A2 셀 값: '{a2_value}'")    
                    sample_template = str(a2_value)
                    sample_template = sample_template.replace('\r\n', '\n').replace('\r', '\n')
                    
                    print(f"A2 셀에서 읽은 템플릿(띄어쓰기 확인): '{sample_template}'")
                    print(f"템플릿 문자 코드 확인: {[ord(c) for c in sample_template[:20]]}")
                    
                    # 가독성 향상을 위한 줄바꿈 처리
                    if "\n" not in sample_template:
                        if "]" in sample_template:
                            bracket_pos = sample_template.find("]")
                            if bracket_pos > 0:
                                sample_template = sample_template[:bracket_pos+1] + "\n" + sample_template[bracket_pos+1:]
                        
                        if "님," in sample_template:
                            sample_template = sample_template.replace("님,", "님,\n")
                        
                        if "쇼핑몰입니다" in sample_template:
                            sample_template = sample_template.replace("쇼핑몰입니다", "쇼핑몰입니다.\n")
                        
                        if "발송됩니다" in sample_template:
                            sample_template = sample_template.replace("발송됩니다", "발송됩니다.\n")
                        
                        if "◎" in sample_template:
                            sample_template = sample_template.replace("◎", "\n◎")
                        
                        if "감사합니다" in sample_template:
                            sample_template = sample_template.replace("감사합니다", "\n감사합니다")
                    
                    while "\n\n" in sample_template: 
                        sample_template = sample_template.replace("\n\n", "\n")
                    sample_template = sample_template.replace(":", ": ")
                    while "  " in sample_template: 
                        sample_template = sample_template.replace("  ", " ")
                    
                    print(f"처리된 템플릿:\n{sample_template}")
                    print(f"처리된 템플릿 문자 코드: {[ord(c) for c in sample_template[:20]]}")
                    has_template_from_a2 = True
                else:
                    print("A2 셀이 비어 있거나 값이 없습니다. 기본 템플릿을 사용합니다.")
            else:
                print("템플릿 시트를 찾을 수 없습니다. 첫 번째 시트의 A2 셀을 확인합니다.")
                try:
                    first_sheet = wb[wb.sheetnames[0]]
                    a2_value = first_sheet['A2'].value
                    if a2_value and a2_value.strip():
                        print(f"첫 번째 시트의 A2 셀 값: '{a2_value}'")
                        sample_template = str(a2_value)
                        has_template_from_a2 = True
                except Exception as e:
                    print(f"첫 번째 시트 A2 셀 접근 중 오류: {str(e)}")
                    print("기본 템플릿을 사용합니다.")
        except Exception as e:
            print(f"템플릿 추출 중 오류 발생: {str(e)}")
            traceback.print_exc()
            print("기본 템플릿을 사용합니다.")
            
        # 데이터 시트 읽기
        print("엑셀 파일 데이터 시트 읽기 시작")
        try:
            # 다양한 시트 이름 시도
            df = None
            sheet_names_tried = []
            
            try:
                # 먼저 모든 시트 이름 가져오기
                all_sheets = pd.ExcelFile(temp_excel_path).sheet_names
                print(f"파일의 모든 시트: {all_sheets}")
                
                # 시트 이름 리스트 준비 - 알려진 이름 + 파일의 모든 시트
                sheet_name_candidates = ['data', 'Data', '데이터', 'Sheet1', '발송', '발송목록', 'data-sheet'] + all_sheets
                
                # 중복 제거
                sheet_name_candidates = list(dict.fromkeys(sheet_name_candidates))
                
                for sheet_name in sheet_name_candidates:
                    try:
                        sheet_names_tried.append(sheet_name)
                        df = pd.read_excel(temp_excel_path, sheet_name=sheet_name, engine='openpyxl')
                        print(f"시트 '{sheet_name}' 발견됨, 행 수: {len(df)}")
                        
                        # 데이터 확인 (최소 헤더 + 1행)
                        if len(df) >= 1:
                            break
                        else:
                            print(f"시트 '{sheet_name}'에 데이터가 없음, 다음 시트 시도")
                            df = None
                    except Exception as e:
                        print(f"시트 '{sheet_name}' 읽기 실패: {str(e)}")
                        continue
            except Exception as e:
                print(f"시트 목록 가져오기 실패: {str(e)}")
            
            # 아직 데이터를 찾지 못했으면 첫 번째 시트 시도
            if df is None:
                try:
                    print("시트 이름 지정 없이 첫 번째 시트 시도")
                    df = pd.read_excel(temp_excel_path, engine='openpyxl')
                    print(f"첫 번째 시트 사용: 행 수: {len(df)}")
                except Exception as e:
                    print(f"첫 번째 시트 읽기 실패: {str(e)}")
            
            # 여전히 데이터가 없으면 오류 반환
            if df is None:
                error_message = f"엑셀 파일에서 데이터를 찾을 수 없습니다. 시도한 시트: {', '.join(sheet_names_tried)}"
                print(error_message)
                return {
                    'success': False,
                    'message': error_message
                }
                
            print(f"데이터 로드 완료: {df.shape[0]}행, {df.shape[1]}열")
            print(f"엑셀 파일 열: {df.columns.tolist()}")
            
            # 체크박스 열 확인
            has_checkbox = False
            checkbox_col = None
            
            # '조건' 열 확인
            for col in df.columns:
                col_str = str(col).lower()
                if col_str == '조건' or col_str == '발송여부' or col_str == 'send' or col_str == '전송':  # '조건' 열이 TRUE/FALSE 값을 가짐
                    has_checkbox = True
                    checkbox_col = col
                    print(f"조건 열 발견: {checkbox_col}")
                    break
            
            # 이름이 없는 첫 번째 열이 체크박스일 수 있음
            if checkbox_col is None:
                for col in df.columns:
                    if str(col).startswith('Unnamed') or col == 0:  # 체크박스는 보통 이름이 없는 첫 번째 열
                        has_checkbox = True
                        checkbox_col = col
                        print(f"체크박스 열 발견: {checkbox_col}")
                        break
                        
            # 필수 열 확인 및 찾기 - 열 이름을 더 유연하게 인식
            name_col = None
            phone_col = None
            date_col = None
            amount_col = None
            product_col = None
            
            # 이름 열 찾기 - 다양한 열 이름 지원
            for col in df.columns:
                col_str = str(col).lower()
                if col_str == '이름' or '성명' in col_str or '고객' in col_str or '고객명' in col_str or '수신자' in col_str or 'name' in col_str:
                    name_col = col
                    print(f"이름 열 발견: {name_col}")
                    break
                    
            # 전화번호 열 찾기 - 다양한 열 이름 지원
            if '휴대폰번호' in df.columns:
                phone_col = '휴대폰번호'
                print(f"정확한 휴대폰번호 열 발견: {phone_col}")
            else:
                for col in df.columns:
                    col_str = str(col).lower()
                    if (col_str == '전화번호' or '수신' in col_str or '휴대' in col_str or '전화' in col_str 
                        or '연락' in col_str or '폰' in col_str or '번호' in col_str or 'phone' in col_str 
                        or 'mobile' in col_str or 'tel' in col_str):
                        phone_col = col
                        print(f"유사한 휴대폰번호 열 발견: {phone_col}")
                        break
                
            # 주문일자 열 찾기 - 다양한 열 이름 지원
            for col in df.columns:
                col_str = str(col).lower()
                if ('주문일자' in col_str or '주문날짜' in col_str or '결제일' in col_str or '주문일' in col_str 
                    or '결제일자' in col_str or '구매일' in col_str or 'order date' in col_str 
                    or 'orderdate' in col_str or 'date' == col_str):
                    date_col = col
                    print(f"주문일자 열 발견: {date_col}")
                    break
                    
            # 주문금액 열 찾기 - 다양한 열 이름 지원
            for col in df.columns:
                col_str = str(col).lower()
                if ('주문금액' in col_str or '결제금액' in col_str or '금액' in col_str or '가격' in col_str 
                    or '비용' in col_str or 'price' in col_str or 'amount' in col_str or 'cost' in col_str):
                    amount_col = col
                    print(f"주문금액 열 발견: {amount_col}")
                    break
                    
            # 주문상품 열 찾기 - 다양한 열 이름 지원
            for col in df.columns:
                col_str = str(col).lower()
                if ('주문상품' in col_str or '상품명' in col_str or '제품명' in col_str or '상품' in col_str 
                    or '제품' in col_str or 'product' in col_str or 'item' in col_str):
                    product_col = col
                    print(f"주문상품 열 발견: {product_col}")
                    break
            
            # 필수 열 확인 - 필수는 아니지만 추천하는 열 표시
            missing_columns = []
            if not name_col: missing_columns.append('이름')
            if not phone_col: missing_columns.append('휴대폰번호')
            
            # 전화번호 열은 필수
            if not phone_col:
                return {
                    'success': False,
                    'message': f'엑셀 파일에 휴대폰번호 열이 없습니다. 수신자 정보를 확인해주세요.'
                }
            
            # 템플릿에 변수가 있는지 확인하고 안내
            import re
            variables = re.findall(r'\{\{([^{}]+)\}\}', sample_template)
            if variables:
                print(f"템플릿에서 발견된 변수: {variables}")
                missing_vars = []
                
                var_mapping = {
                    '이름': name_col,
                    '주문일자': date_col,
                    '주문금액': amount_col,
                    '주문상품': product_col
                }
                
                for var in variables:
                    if var in var_mapping and not var_mapping[var]:
                        missing_vars.append(var)
                
                if missing_vars:
                    print(f"[경고] 템플릿에 사용된 변수 중 해당하는 열을 찾을 수 없습니다: {', '.join(missing_vars)}")
                    print("해당 변수는 치환되지 않거나 기본값으로 대체될 수 있습니다.")
            
            # 체크된 행만 필터링
            if has_checkbox:
                checkbox_values = df[checkbox_col].astype(str).str.upper()
                is_checked = checkbox_values.isin(['TRUE', '1', 'YES', 'Y', 'O', 'V', 'T', 'TRUE', 'OK'])
                filtered_df = df[is_checked].copy()
                print(f"체크된 행만 필터링: 전체 {df.shape[0]}행 중 {filtered_df.shape[0]}행 선택됨")
            else:
                filtered_df = df.copy()
                print(f"체크박스 없음, 모든 행 처리: {filtered_df.shape[0]}행")
            
            if filtered_df.empty:
                return {
                    'success': False,
                    'message': '처리할 행이 없습니다. 체크박스가 있는 경우 최소 하나의 행을 체크하세요.'
                }
            
            # 미리보기용 메시지 생성
            preview_messages = []
            recipients = []
            
            processed_count = 0
            skipped_count = 0
            
            # 각 행 처리
            for idx, row in filtered_df.iterrows():
                # 전화번호 확인
                if phone_col not in row.index or pd.isna(row[phone_col]) or str(row[phone_col]).strip() == '':
                    print(f"[경고] 행 {idx+1}: 수신번호가 없음 -> 건너뛰기")
                    skipped_count += 1
                    continue
                
                # 전화번호 정제
                phone = str(row[phone_col])
                phone = ''.join(filter(str.isdigit, phone))
                
                if not phone or len(phone) < 10:
                    print(f"[경고] 행 {idx+1}: 유효하지 않은 전화번호: {row[phone_col]} -> 건너뛰기")
                    skipped_count += 1
                    continue
                
                # 변수값 추출
                name = str(row[name_col]) if name_col and not pd.isna(row[name_col]) else ""
                date = str(row[date_col]) if date_col and not pd.isna(row[date_col]) else ""
                amount = str(row[amount_col]) if amount_col and not pd.isna(row[amount_col]) else ""
                product = str(row[product_col]) if product_col and not pd.isna(row[product_col]) else ""
                
                # 날짜 형식 정리 - 더 많은 형식 지원
                if date and (date_col is not None) and ('일자' in str(date_col) or '날짜' in str(date_col) or 'date' in str(date_col).lower()):
                    try:
                        # 여러 날짜 형식 처리
                        if ' 00:00:00' in date:  # 'YYYY-MM-DD 00:00:00' 형식
                            date = date.split(' ')[0]
                        elif 'T00:00:00' in date:  # 'YYYY-MM-DDT00:00:00' 형식
                            date = date.split('T')[0]
                        elif '.' in date and len(date) > 10:  # 'YYYY-MM-DD.000000' 형식
                            date = date.split('.')[0]
                        
                        # pandas Timestamp 처리
                        from pandas import Timestamp
                        if isinstance(row[date_col], Timestamp):
                            date = row[date_col].strftime('%Y-%m-%d')
                            
                        print(f"날짜 변환: {row[date_col]} -> {date}")
                    except Exception as e:
                        print(f"날짜 변환 중 오류: {str(e)}")
                
                # 금액 형식 정리 - 천 단위 구분기호 처리
                if amount and amount_col is not None:
                    try:
                        # 원본 값 백업
                        original_amount = amount
                        # 숫자만 추출
                        numeric_amount = ''.join(filter(str.isdigit, amount))
                        if numeric_amount:
                            # 천 단위 구분기호 추가
                            formatted_amount = format(int(numeric_amount), ',')
                            
                            # '원'이 있는지 확인하고 유지
                            if '원' in original_amount:
                                formatted_amount += '원'
                                
                            amount = formatted_amount
                            print(f"금액 변환: {row[amount_col]} -> {amount}")
                    except Exception as e:
                        print(f"금액 변환 중 오류: {str(e)}")
                
                # 템플릿에 변수 적용 - 변수 치환 기능 강화
                message_text = sample_template
                
                # 변수 치환 - 이중 중괄호 처리 개선
                import re
                
                # 정해진 변수들 처리
                var_dict = {
                    '이름': name,
                    '주문일자': date,
                    '주문금액': amount,
                    '주문상품': product
                }
                
                # 변수 치환 처리
                for var_name, var_value in var_dict.items():
                    if var_value:  # 값이 있을 때만 치환
                        pattern = re.compile(r'\{\{' + re.escape(var_name) + r'\}\}')
                        if re.search(pattern, message_text):
                            message_text = re.sub(pattern, var_value, message_text)
                            print(f"{var_name} 치환: '{var_value}'")
                
                # 동적 변수 치환 - 열 이름 기반
                remaining_vars = re.findall(r'\{\{([^{}]+)\}\}', message_text)
                for var_name in remaining_vars:
                    if var_name in row.index and not pd.isna(row[var_name]):
                        var_value = str(row[var_name])
                        
                        # 날짜 형식 특별 처리
                        if '일자' in var_name or '날짜' in var_name or 'date' in var_name.lower():
                            try:
                                if ' 00:00:00' in var_value:
                                    var_value = var_value.split(' ')[0]
                                elif 'T00:00:00' in var_value:
                                    var_value = var_value.split('T')[0]
                                
                                from pandas import Timestamp
                                if isinstance(row[var_name], Timestamp):
                                    var_value = row[var_name].strftime('%Y-%m-%d')
                            except Exception as e:
                                print(f"변수 {var_name} 날짜 변환 중 오류: {str(e)}")
                        
                        # 금액 형식 특별 처리
                        if '금액' in var_name or '가격' in var_name or 'price' in var_name.lower() or 'amount' in var_name.lower():
                            try:
                                # 원본 값 백업
                                original_value = var_value
                                numeric_value = ''.join(filter(str.isdigit, var_value))
                                if numeric_value:
                                    formatted_value = format(int(numeric_value), ',')
                                    
                                    # '원'이 있는지 확인하고 유지
                                    if '원' in original_value:
                                        formatted_value += '원'
                                        
                                    var_value = formatted_value
                            except Exception as e:
                                print(f"변수 {var_name} 금액 변환 중 오류: {str(e)}")
                                
                        var_pattern = re.compile(r'\{\{' + re.escape(var_name) + r'\}\}')
                        message_text = re.sub(var_pattern, var_value, message_text)
                        print(f"변수 {var_name} 치환: '{var_value}'")
                        
                # 남은 이중 중괄호 제거
                if '{{' in message_text and '}}' in message_text:
                    message_text = re.sub(r'\{\{([^{}]+)\}\}', r'\1', message_text)
                
                # 메시지 태그 또는 자동문자 관련 ID 추가
                message_id = f"AUTO_{idx+1}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
                
                # 메시지 객체 생성
                message = {
                    'to': phone.replace('-', ''),  # 하이픈 제거
                    'from': sender_phone.replace('-', ''),  # 하이픈 제거
                    'text': format_message_for_sms(message_text),  # 메시지 포맷팅 적용
                    'type': 'LMS' if len(message_text) > 90 else 'SMS'  # 명시적으로 메시지 타입 지정
                }
                
                recipients.append(message)
                
                # 미리보기용 메시지 추가
                preview = {
                    'index': len(preview_messages) + 1,
                    'phone': phone,
                    'text': message_text
                }
                preview_messages.append(preview)
                processed_count += 1
                
                # 첫 몇 개 메시지는 상세 로그 기록
                if processed_count <= 3:
                    preview_text = message_text[:50] + "..." if len(message_text) > 50 else message_text
                    print(f"행 {idx+1}: 메시지 생성 (길이: {len(message_text)}자) - {preview_text}")
            
            print(f"\n총 {filtered_df.shape[0]}행 중 {processed_count}개 처리됨, {skipped_count}개 건너뜀")
            
            if has_checkbox:
                print("체크박스 선택된 항목만 처리되었습니다.")
            
            # 결과 반환
            result = {
                'success': True,
                'message': f'자동 메시지 템플릿 처리 완료: {processed_count}건',
                'total': processed_count,
                'preview': preview_messages[:5],  # 미리보기는 최대 5건만 표시
                'recipients': recipients  # 발송 목적인 경우 사용할 전체 수신자 목록
            }
            
            return result
        except Exception as e:
            print(f"데이터 처리 중 오류 발생: {str(e)}")
            traceback.print_exc()
            return {
                'success': False,
                'message': f'엑셀 파일 처리 중 오류가 발생했습니다: {str(e)}'
            }
    except Exception as e:
        print(f"자동 메시지 처리 중 오류 발생: {str(e)}")
        traceback.print_exc()
        return {
            'success': False,
            'message': f'자동 메시지 처리 중 오류가 발생했습니다: {str(e)}'
        }